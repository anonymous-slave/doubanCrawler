#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')

hds = [{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0;
    book_list = []
    try_times = 0

    while(1):
        url = 'https://book.douban.com/tag/'+urllib.quote(book_tag)+'?start='+str(page_num*20)
        # print 'Successflly Connected'
        time.sleep(np.random.rand()*5)

        try:
            req = urllib2.Request(url,headers=hds[page_num%len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
        except (urllib2.HTTPError, urllib2.URLError),e:
            print e
            continue

        soup = BeautifulSoup(plain_text,"lxml")
        list_soup = soup.find('div',{'id':'subject_list'})
        items = list_soup.find('li',{'class':'subject-item'})
        # print list_soup

        try_times += 1;
        if items == None and try_times < 5:
            continue
        elif items == None or len(items) <= 1:
            break

        for book_info in list_soup.findAll('div',{'class':'info'}):
            # print book_info
            title = book_info.find('a').get('title')
            desc = book_info.find('div',{'class':'pub'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a').get('href')
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者: 暂无'
            try:
                pub_info = '出版信息： '+ '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息：暂无'
            try:
                rating = book_info.find('span',{'class':'rating_nums'}).string.strip()
            except:
                rating = '0.0'
            try:
                people_num = book_info.find('span',{'class':'pl'}).string.strip()
            except:
                people_num = '0'

            book_list.append([title,rating,people_num,author_info,pub_info,book_url])
            try_times = 0
        page_num += 1
        print 'Downloading Information From Page %d' %page_num
    return book_list

def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社','链接'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],bl[1],bl[2],bl[3],bl[4],bl[5]])
            count += 1
        save_path = 'book_list_myself'
        for i in range(len(book_tag_lists)):
            save_path += ('-'+book_tag_lists[i].decode())
        save_path += '.xlsx'
        wb.save(save_path)


if __name__=='__main__':
    book_tag_lists = ['名著']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)